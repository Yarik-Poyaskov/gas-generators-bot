from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from typing import List, Any
import json
import io
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from app.api.deps import get_current_user, require_role
from app.api.models import (
    ObjectInfo, ReportInfo, TraderPublishRequest, 
    TraderObjectSchedule, ExportRequest
)
from app.db.database import (
    get_objects_with_latest_status, get_all_objects,
    get_recent_reports, get_reports_by_range,
    add_trader_schedule, add_trader_announcement
)
from app.config import config

router = APIRouter(prefix="/data", tags=["Data"])

@router.get("/objects", response_model=List[ObjectInfo])
async def get_objects(current_user: dict = Depends(get_current_user)):
    """Returns all objects with their current status for the dashboard."""
    objects = await get_objects_with_latest_status()
    
    # Filter if user is not admin or trader (regular users see only their objects)
    if current_user["role"] == "user":
        from app.db.database import get_user_objects_by_tg_id
        user_objs = await get_user_objects_by_tg_id(current_user["user_id"])
        allowed_ids = {o["id"] for o in user_objs}
        objects = [o for o in objects if o["id"] in allowed_ids]
        
    return objects

@router.get("/reports", response_model=List[ReportInfo])
async def get_reports(limit: int = 50, offset: int = 0, current_user: dict = Depends(get_current_user)):
    """Returns the history of reports with pagination."""
    reports = await get_recent_reports(limit=limit, offset=offset)
    
    # Filter for regular users: they can only see reports for their objects
    if current_user["role"] == "user":
        from app.db.database import get_user_objects_by_tg_id
        user_objs = await get_user_objects_by_tg_id(current_user["user_id"])
        allowed_names = {o["name"] for o in user_objs}
        # Note: report has tc_name which contains object name
        filtered_reports = []
        for r in reports:
            for name in allowed_names:
                if name in r["tc_name"]:
                    filtered_reports.append(r)
                    break
        return filtered_reports
        
    return reports

@router.post("/trader/parse")
async def parse_schedule_text(payload: dict, current_user: dict = Depends(require_role(["admin", "trader"]))):
    """Parses trader's schedule text using the existing bot logic."""
    text = payload.get("text")
    if not text:
        raise HTTPException(status_code=400, detail="Missing text")
        
    from app.handlers.trader_parser import parse_trader_message, load_mapping
    mapping = load_mapping()
    parsed_data, date_str = parse_trader_message(text, mapping)
    
    if not parsed_data:
        return {"success": False, "error": date_str or "Could not parse text"}
        
    return {
        "success": True, 
        "date": date_str,
        "data": parsed_data
    }

@router.post("/trader/publish")
async def publish_schedule(payload: TraderPublishRequest, current_user: dict = Depends(require_role(["admin", "trader"]))):
    """Saves and publishes schedules to Telegram groups from the Web interface."""
    from app.api.auth import bot_instance
    if not bot_instance:
        raise HTTPException(status_code=500, detail="Bot instance not initialized")
        
    from app.keyboards.inline import get_schedule_confirm_kb
    from fastapi.encoders import jsonable_encoder
    
    all_objs = await get_all_objects()
    obj_map = {o['name']: (o['id'], o['telegram_group_id']) for o in all_objs}
    
    success_count = 0
    published_items = []
    target_date_db = ""

    for item in payload.items:
        # intervals are list of objects, need to serialize to JSON for DB
        intervals_json = json.dumps(jsonable_encoder(item.intervals), ensure_ascii=False)
        target_date_db = item.target_date # YYYY-MM-DD
        
        # Match DB name
        obj_info = next((info for name, info in obj_map.items() if item.db_name in name), None)
        if obj_info:
            obj_id, group_id = obj_info
            # Save to DB
            sched_id = await add_trader_schedule(
                obj_id, 
                current_user["user_id"], 
                item.target_date, 
                intervals_json, 
                item.is_not_working
            )
            success_count += 1
            published_items.append(item)
            
            # Send to Object Group
            if group_id:
                try:
                    date_disp = datetime.strptime(item.target_date, "%Y-%m-%d").strftime("%d.%m.%Y")
                    msg = f"📅 <b>ГРАФІК РОБОТИ: {date_disp}</b>\n🏢 Об'єкт: <b>{item.db_name}</b>\n\n"
                    if item.is_not_working:
                        msg += "❌ <b>ГПУ НЕ ПРАЦЮЄ</b>"
                    else:
                        msg += "\n".join([f"• {i.start} - {i.end} | {i.power}% | {i.mode}" for i in item.intervals])
                    
                    sent_announcement = await bot_instance.send_message(
                        chat_id=group_id, 
                        text=msg + "\n\n👉 Підтвердіть графік.", 
                        reply_markup=get_schedule_confirm_kb(sched_id), 
                        parse_mode="HTML"
                    )
                    
                    # Track for revocation
                    await add_trader_announcement(
                        trader_id=current_user["user_id"], 
                        target_date=item.target_date, 
                        chat_id=group_id, 
                        message_id=sent_announcement.message_id,
                        object_id=obj_id,
                        message_type='announcement'
                    )
                except Exception as e:
                    print(f"Error sending to group {group_id}: {e}")

    # Send summary to Trader Group
    if success_count > 0:
        from app.handlers.trader_parser import format_review_text
        from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
        
        # format_review_text expects list of dicts
        review_data = jsonable_encoder(published_items)
        date_disp = datetime.strptime(target_date_db, "%Y-%m-%d").strftime("%d/%m/%Y")
        
        summary_text = f"🚀 <b>Графік опубліковано через Web-портал</b>\n"
        summary_text += f"👤 Користувач: <b>{current_user['full_name']}</b>\n\n"
        summary_text += format_review_text(review_data, date_disp)
        summary_text += f"\n✅ <b>Збережено об'єктів: {success_count}</b>"
        
        revoke_kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🗑 Відкликати графік", callback_data=f"revoke_sched:{target_date_db}")]
        ])
        
        try:
            await bot_instance.send_message(
                chat_id=config.trader_monitor_group_id,
                text=summary_text,
                reply_markup=revoke_kb,
                parse_mode="HTML"
            )
        except Exception as e:
            print(f"Error sending summary to trader group: {e}")

    return {"success": True, "count": success_count}

@router.get("/reports/export")
async def export_reports(start_date: str, end_date: str, current_user: dict = Depends(require_role(["admin"]))):
    """Generates and returns an Excel file with reports for the specified period."""
    reports = await get_reports_by_range(start_date, end_date)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Звіти ГПУ"
    
    # Define headers
    headers = [
        "ID", "Дата/Час", "Об'єкт", "Заповнив", "Режим", "Час", 
        "Потужність (%)", "Потужність (кВт)", "Статус ГПУ", 
        "Напруга АКБ", "Тиск (До)", "Тиск (Після)", 
        "МВт*год (всього)", "Мотогодини", "Олива (ліміт)", "Тип часу"
    ]
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        
    # Add data
    for row_idx, r in enumerate(reports, 2):
        ws.cell(row=row_idx, column=1, value=r["id"])
        ws.cell(row=row_idx, column=2, value=r["created_at"])
        ws.cell(row=row_idx, column=3, value=re.sub(r"[\(\)]", "", r["tc_name"]))
        ws.cell(row=row_idx, column=4, value=r["full_name"])
        ws.cell(row=row_idx, column=5, value=r["work_mode"])
        ws.cell(row=row_idx, column=6, value=r["start_time"])
        ws.cell(row=row_idx, column=7, value=r["load_power_percent"])
        ws.cell(row=row_idx, column=8, value=r["load_power_kw"])
        ws.cell(row=row_idx, column=9, value=r["gpu_status"])
        ws.cell(row=row_idx, column=10, value=r["battery_voltage"])
        ws.cell(row=row_idx, column=11, value=r["pressure_before"])
        ws.cell(row=row_idx, column=12, value=r["pressure_after"])
        ws.cell(row=row_idx, column=13, value=r["total_mwh"])
        ws.cell(row=row_idx, column=14, value=r["total_hours"])
        ws.cell(row=row_idx, column=15, value=r["oil_sampling_limit"])
        ws.cell(row=row_idx, column=16, value=r["time_type"])

    # Auto-adjust column width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"GPU_Reports_{start_date}_{end_date}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
